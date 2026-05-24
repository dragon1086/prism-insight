"""Contract tests: verify mock_kis_server.py responses match the KIS Open API spec.

Spec source: the KIS Open API Excel file (~339 sheets). The file lives outside
the repo (typically alongside it). Search order:

1. `KIS_API_SPEC_XLSX` env var — explicit absolute path
2. `../*KIS*API*.xlsx` relative to the repo root — common shipping location

If neither resolves the test is skipped. openpyxl is also required and the
suite is skipped (not errored) when missing — these tests document expected
behavior but should never block CI when the spec ships out-of-band.

The check is intentionally **shape only**: we assert every spec Response Body
field is present in the mock response. Field types and value validation are
not checked — KIS frequently overloads numeric fields as strings, so the
contract is "key exists".
"""
from __future__ import annotations

import glob
import os
import sys
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

pytest.importorskip("fastapi", reason="fastapi not installed — pip install fastapi")
openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed — pip install openpyxl")

from fastapi.testclient import TestClient  # noqa: E402

from tests import mock_kis_server as mks  # noqa: E402


def _resolve_xlsx_path() -> str | None:
    """Find the KIS Open API spec XLSX.

    Resolution order:
      1. `KIS_API_SPEC_XLSX` env var — explicit absolute path
      2. Walk up the directory tree from this file looking for any
         *API*.xlsx (handles both regular checkouts and git worktrees,
         which sit several levels deeper than the repo root).
    """
    explicit = os.environ.get("KIS_API_SPEC_XLSX")
    if explicit and Path(explicit).is_file():
        return explicit
    # Walk up up to 8 levels (regular checkout = 1, worktree = ~5).
    here = Path(__file__).resolve()
    for parent in [*here.parents][:8]:
        for pattern in ("*KIS*API*.xlsx", "*API*.xlsx"):
            for candidate in sorted(glob.glob(str(parent / pattern))):
                if Path(candidate).is_file():
                    return candidate
    return None


XLSX_PATH = _resolve_xlsx_path()


# Endpoint catalog: (sheet name, HTTP method, URL, request payload, layout).
# layout describes how the response body maps to spec sections:
#   - "flat"   : top-level keys are the spec body fields (token endpoints)
#   - "single" : body has rt_cd/msg* envelope plus a single `output` object
#   - "paired" : body has `output1` (array) and `output2` (array) sections
#   - "array"  : body has a single `output` array; per-row spec applies
_ENDPOINTS = [
    {
        "name": "주식주문(현금)",
        "sheet": "주식주문(현금)",
        "method": "POST",
        "url": "/uapi/domestic-stock/v1/trading/order-cash",
        "json": {"CANO": "50000000", "ACNT_PRDT_CD": "01", "PDNO": "005930",
                  "ORD_DVSN": "01", "ORD_QTY": "1", "ORD_UNPR": "0"},
        "headers": {"tr_id": "VTTC0012U"},
        "layout": "single",
    },
    {
        "name": "주식예약주문",
        "sheet": "주식예약주문",
        "method": "POST",
        "url": "/uapi/domestic-stock/v1/trading/order-resv",
        "json": {"CANO": "50000000", "ACNT_PRDT_CD": "01", "PDNO": "005930",
                  "ORD_QTY": "1", "ORD_UNPR": "1000",
                  "SLL_BUY_DVSN_CD": "02", "ORD_DVSN_CD": "00",
                  "ORD_OBJT_CBLC_DVSN_CD": "10"},
        "headers": {"tr_id": "CTSC0008U"},
        "layout": "single",
    },
    {
        "name": "주식잔고조회",
        "sheet": "주식잔고조회",
        "method": "GET",
        "url": "/uapi/domestic-stock/v1/trading/inquire-balance",
        "params": {"CANO": "50000000", "ACNT_PRDT_CD": "01"},
        "layout": "paired",
        "setup": "seed_holding",
    },
    {
        "name": "매수가능조회",
        "sheet": "매수가능조회",
        "method": "GET",
        "url": "/uapi/domestic-stock/v1/trading/inquire-psbl-order",
        "params": {"CANO": "50000000", "ACNT_PRDT_CD": "01",
                    "PDNO": "005930", "ORD_UNPR": "70000"},
        "layout": "single",
    },
    {
        "name": "주식일별주문체결조회",
        "sheet": "주식일별주문체결조회",
        "method": "GET",
        "url": "/uapi/domestic-stock/v1/trading/inquire-daily-ccld",
        "params": {"CANO": "50000000", "ACNT_PRDT_CD": "01"},
        "layout": "paired",
        "setup": "place_order",
    },
    {
        "name": "주식현재가 시세",
        "sheet": "주식현재가 시세",
        "method": "GET",
        "url": "/uapi/domestic-stock/v1/quotations/inquire-price",
        "params": {"fid_cond_mrkt_div_code": "J", "fid_input_iscd": "005930"},
        "layout": "single",
    },
    {
        "name": "주식현재가 일자별",
        "sheet": "주식현재가 일자별",
        "method": "GET",
        "url": "/uapi/domestic-stock/v1/quotations/inquire-daily-price",
        "params": {"fid_cond_mrkt_div_code": "J", "fid_input_iscd": "005930",
                    "fid_period_div_code": "D", "fid_org_adj_prc": "0"},
        "layout": "array",
    },
    {
        "name": "접근토큰발급(P)",
        "sheet": "접근토큰발급(P)",
        "method": "POST",
        "url": "/oauth2/tokenP",
        "json": {"grant_type": "client_credentials", "appkey": "k", "appsecret": "s"},
        "layout": "flat",
    },
]


def _extract_response_body_sections(sheet) -> dict[str, list[str]]:
    """Walk the spec sheet and collect Response Body field names per section.

    The Excel layout is consistent across all KIS endpoints:
        Row N:       "Response Body" | rt_cd | …
        Row N+k:     ... msg_cd, msg1 (envelope fields, section='body')
        Row M:       output | object  → new section 'output'
        Row M+k:     ... field rows (section='output')
        (optionally another output2 array)
        Row L:       "Response Example" → stop
    """
    sections: dict[str, list[str]] = {}
    current: str | None = None
    in_rb = False
    for row in sheet.iter_rows(values_only=True):
        vals = [str(c) if c is not None else "" for c in row]
        first = vals[0] if vals else ""
        second = vals[1] if len(vals) > 1 else ""
        type_ = vals[3] if len(vals) > 3 else ""

        if first == "Response Body":
            in_rb = True
            current = "body"
            sections.setdefault("body", [])
            continue
        if first in ("Response Example", "Example", "Request Header", "Response Header",
                      "Layout", "Request Body", "Query Parameter"):
            in_rb = False
            continue
        if not in_rb or not second:
            continue
        if type_ in ("object", "object array"):
            current = second
            sections.setdefault(current, [])
            continue
        sections.setdefault(current or "body", []).append(second)
    return sections


@pytest.fixture(scope="module")
def spec_workbook():
    if XLSX_PATH is None:
        pytest.skip(
            "KIS API spec XLSX not found. Set KIS_API_SPEC_XLSX or place a "
            "*KIS*API*.xlsx file alongside the repo root."
        )
    from openpyxl import load_workbook
    return load_workbook(XLSX_PATH, read_only=True, data_only=True)


@pytest.fixture
def client():
    mks.STATE.reset()
    return TestClient(mks.app)


def _apply_setup(client, setup: str | None):
    if setup == "seed_holding":
        client.post("/__mock__/seed_holding", json={
            "CANO": "50000000", "ACNT_PRDT_CD": "01",
            "PDNO": "005930", "qty": 10, "avg_price": 70000,
        })
    elif setup == "place_order":
        client.post("/uapi/domestic-stock/v1/trading/order-cash",
                    json={"CANO": "50000000", "ACNT_PRDT_CD": "01",
                          "PDNO": "005930", "ORD_DVSN": "01",
                          "ORD_QTY": "3", "ORD_UNPR": "0"},
                    headers={"tr_id": "VTTC0012U"})


@pytest.mark.parametrize("ep", _ENDPOINTS, ids=[e["name"] for e in _ENDPOINTS])
def test_endpoint_response_matches_spec(ep, spec_workbook, client):
    """For each documented KIS endpoint, the mock response must contain every
    Response Body field the spec defines."""
    sheet = spec_workbook[ep["sheet"]]
    spec_sections = _extract_response_body_sections(sheet)

    _apply_setup(client, ep.get("setup"))

    if ep["method"] == "POST":
        resp = client.post(ep["url"], json=ep.get("json", {}),
                           headers=ep.get("headers", {}))
    else:
        resp = client.get(ep["url"], params=ep.get("params", {}))

    assert resp.status_code == 200, f"HTTP {resp.status_code}: {resp.text[:200]}"
    body = resp.json()

    gaps: list[str] = []
    layout = ep["layout"]

    if layout == "flat":
        spec_fields = set(spec_sections.get("body", []))
        actual = set(body.keys())
        missing = spec_fields - actual
        if missing:
            gaps.append(f"top-level missing: {sorted(missing)}")

    elif layout == "single":
        spec_body = set(spec_sections.get("body", []))
        body_missing = (spec_body - {"output"}) - set(body.keys())
        if body_missing:
            gaps.append(f"envelope missing: {sorted(body_missing)}")
        out = body.get("output", {})
        spec_out = set(spec_sections.get("output", []))
        actual_out = set(out.keys()) if isinstance(out, dict) else set()
        out_missing = spec_out - actual_out
        if out_missing:
            gaps.append(f"output missing: {sorted(out_missing)}")

    elif layout == "paired":
        spec_body = set(spec_sections.get("body", []))
        body_missing = (spec_body - {"output1", "output2"}) - set(body.keys())
        if body_missing:
            gaps.append(f"envelope missing: {sorted(body_missing)}")
        for sec in ("output1", "output2"):
            spec_fields = set(spec_sections.get(sec, []))
            rows = body.get(sec, [])
            if not rows:
                if spec_fields:
                    gaps.append(f"{sec} empty (cannot verify {len(spec_fields)} fields)")
                continue
            first = rows[0] if isinstance(rows, list) else rows
            actual = set(first.keys()) if isinstance(first, dict) else set()
            missing = spec_fields - actual
            if missing:
                gaps.append(f"{sec} missing: {sorted(missing)}")

    elif layout == "array":
        spec_out = set(spec_sections.get("output", []))
        rows = body.get("output", [])
        assert rows, "expected at least one row in output array"
        first = rows[0]
        actual = set(first.keys()) if isinstance(first, dict) else set()
        missing = spec_out - actual
        if missing:
            gaps.append(f"output row missing: {sorted(missing)}")

    assert not gaps, "spec gaps:\n  " + "\n  ".join(gaps)


def test_spec_xlsx_resolves_in_dev_env():
    """Sanity: confirm the XLSX lookup logic finds something in dev. This is
    informational — when the file is genuinely absent the parametrized tests
    above skip themselves via the fixture, but this gives a clear signal in
    the test output when the contract suite isn't actually running."""
    if XLSX_PATH is None:
        pytest.skip("KIS API spec XLSX not present; contract tests will skip")
    assert Path(XLSX_PATH).is_file()
